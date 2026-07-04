"""Bookings xlsx export service.

Builds a filtered, role-scoped queryset and writes an xlsx file using
openpyxl.  All related data is loaded via ``select_related`` in a single
SQL JOIN — no N+1 queries.  Large exports stream rows via
``QuerySet.iterator(chunk_size=1000)`` to keep worker memory constant.
"""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable

from django.conf import settings
from django.db.models import QuerySet
from django.utils import timezone

from accounts.branch_scope import staff_branch_id
from bookings.models import Booking
from bookings.query_filters import apply_booking_export_filters

logger = logging.getLogger("vasavi.bookings.export")


def booking_export_download_api_path(export_id: str) -> str:
    """Staff-authenticated download route (never expose raw MEDIA URLs to the portal)."""
    return f"/api/v1/staff/bookings/export/{export_id}/download/"


def user_can_access_booking_export(export, user) -> bool:
    if user.role == "super_admin":
        return True
    return export.requested_by_id == user.pk


def export_is_downloadable(export) -> bool:
    from bookings.models import BookingExport

    if export.status != BookingExport.Status.READY:
        return False
    if export.expires_at and export.expires_at <= timezone.now():
        return False
    if not export.file_path:
        return False
    return Path(export.file_path).is_file()


def serialize_booking_export(export, *, include_filters: bool = True) -> dict[str, Any]:
    """JSON payload for list/status endpoints."""
    download_url = None
    if export_is_downloadable(export):
        download_url = booking_export_download_api_path(str(export.pk))

    filters = {}
    if include_filters and export.filters_applied:
        filters = {
            k: v
            for k, v in export.filters_applied.items()
            if not str(k).startswith("_") and v not in (None, "", "all")
        }

    branch_name = export.branch.name if getattr(export, "branch", None) else "All branches"

    return {
        "export_id": str(export.pk),
        "status": export.status,
        "progress_percent": export.progress_percent or 0,
        "estimated_count": export.estimated_count,
        "record_count": export.record_count,
        "download_url": download_url,
        "error_message": export.error_message or None,
        "created_at": export.created_at.isoformat() if export.created_at else None,
        "expires_at": export.expires_at.isoformat() if export.expires_at else None,
        "branch_name": branch_name,
        "filters_applied": filters,
    }

# ---------------------------------------------------------------------------
# Column definitions — single source of truth
# ---------------------------------------------------------------------------

EXPORT_HEADERS: list[str] = [
    "Booking ID",
    "Booking Date",
    "Guest Name",
    "Phone Number",
    "Email",
    "Hotel / Branch",
    "City",
    "Room Number",
    "Room Type",
    "Rate per Night (₹)",
    "Check-In Date",
    "Check-Out Date",
    "No. of Nights",
    "No. of Guests",
    "Booking Status",
    "Payment Status",
    "Payment Method",
    "Payment Reference",
    "Base Amount (₹)",
    "Discount (₹)",
    "Total Amount (₹)",
    "Refund Amount (₹)",
    "Cancelled By Role",
    "Cancellation Reason",
    "Notes",
    "Created At",
    "Updated At",
]

# Header background colour (navy, matching Vasavi brand)
HEADER_BG_HEX = "1E3A5F"
HEADER_FG_HEX = "FFFFFF"

# ---------------------------------------------------------------------------
# Queryset builder — bulk, no N+1
# ---------------------------------------------------------------------------


def build_booking_export_queryset(
    filters: dict[str, Any],
    requesting_user,
) -> QuerySet:
    """Return a scoped, filtered queryset with all joins pre-loaded.

    Security contract
    -----------------
    * Branch admin → branch is **always** taken from ``AdminBranch`` assignment,
      never from ``filters``.  Client-supplied ``branch_id`` is silently ignored.
    * Super admin → optional ``branch_id`` and ``city`` filters from ``filters``.
    * All related tables are joined via ``select_related`` in a single SQL query
      to eliminate N+1 when iterating rows.
    """
    qs = (
        Booking.objects.filter(is_deleted=False)
        .select_related(
            "user",             # guest_name / phone / email fallback
            "room",             # room_number, base_price_per_night, capacity
            "room__room_type",  # room type name — avoids second join
            "branch",           # branch name + city
        )
        .order_by("-created_at")
    )

    # -- Server-enforced branch scoping (NEVER trust filters for role=admin) --
    if requesting_user.role == "admin":
        branch_id = staff_branch_id(requesting_user)
        if not branch_id:
            return qs.none()
        qs = qs.filter(branch_id=branch_id)

    elif requesting_user.role == "super_admin":
        # Super admin may optionally scope to one branch or city
        branch_id = (filters.get("branch_id") or "").strip()
        if branch_id:
            qs = qs.filter(branch_id=branch_id)

        city = (filters.get("city") or "").strip()
        if city:
            qs = qs.filter(branch__city__icontains=city)

    # -- Common filters (applied for both roles) ------------------------------
    qs = apply_booking_export_filters(qs, filters)

    return qs


# ---------------------------------------------------------------------------
# Row builder — pure function, called inside iterator loop
# ---------------------------------------------------------------------------


def _fmt_date(d) -> str:
    """Format a date object as DD-MM-YYYY."""
    if d is None:
        return ""
    return d.strftime("%d-%m-%Y")


def _fmt_datetime(dt) -> str:
    """Format a datetime object as DD-MM-YYYY HH:MM."""
    if dt is None:
        return ""
    local_dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return local_dt.strftime("%d-%m-%Y %H:%M")


def _paise_to_inr(paise: int | None) -> str:
    """Convert paise integer to INR string with 2 decimal places."""
    if paise is None:
        return "0.00"
    return f"{paise / 100:.2f}"


def _booking_to_row(b: Booking) -> list[str]:
    """Convert one Booking instance to an ordered list of cell values.

    All related objects (user, room, room__room_type, branch) are already
    joined via select_related — accessing them here costs zero extra queries.
    """
    guest_name  = (b.guest_name  or "").strip() or (b.user.name  or "") if b.user_id else ""
    guest_phone = (b.guest_phone or "").strip() or (b.user.phone or "") if b.user_id else ""
    email       = (b.user.email  or "") if b.user_id else ""

    branch_name = b.branch.name if b.branch_id else ""
    branch_city = b.branch.city if b.branch_id else ""

    room_number    = b.room.room_number if b.room_id else ""
    room_type_name = b.room.room_type.name if b.room_id and b.room.room_type_id else ""
    rate_per_night = _paise_to_inr(b.room.base_price_per_night) if b.room_id else "0.00"

    return [
        b.booking_reference,
        _fmt_datetime(b.created_at),
        guest_name,
        guest_phone,
        email,
        branch_name,
        branch_city,
        room_number,
        room_type_name,
        rate_per_night,
        _fmt_date(b.check_in_date),
        _fmt_date(b.check_out_date),
        str(b.nights),
        str(b.guest_count),
        b.get_status_display(),
        b.get_payment_status_display(),
        b.get_payment_gateway_display() if b.payment_gateway else "",
        b.payment_reference or "",
        _paise_to_inr(b.base_amount),
        _paise_to_inr(b.discount_amount),
        _paise_to_inr(b.final_amount),
        _paise_to_inr(b.refund_amount),
        b.get_cancel_initiated_by_role_display() if b.cancel_initiated_by_role else "",
        b.cancellation_reason or "",
        b.notes or "",
        _fmt_datetime(b.created_at),
        _fmt_datetime(b.updated_at),
    ]


# ---------------------------------------------------------------------------
# xlsx writer
# ---------------------------------------------------------------------------


def build_bookings_xlsx(
    qs: QuerySet,
    file_path: Path,
    *,
    on_progress: Callable[[int, int], None] | None = None,
    progress_total_hint: int | None = None,
) -> int:
    """Write an xlsx export file to *file_path* and return the row count.

    Design decisions
    ----------------
    * ``qs.iterator(chunk_size=1000)`` — streams DB rows in batches of 1 000,
      never materialising the entire queryset in memory.  Safe for 50 k+ rows.
    * ``openpyxl`` write-only mode is NOT used because we need auto column
      widths (requires reading cell values after write).  For 50 k rows the
      in-memory workbook is ~30–50 MB which is acceptable in the Celery worker.
    * Header row is frozen (``freeze_panes``) and styled with navy background.
    * Amount columns use right-alignment; date columns use centre-alignment.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # -- header row -----------------------------------------------------------
    header_fill = PatternFill(
        start_color=HEADER_BG_HEX,
        end_color=HEADER_BG_HEX,
        fill_type="solid",
    )
    header_font = Font(bold=True, color=HEADER_FG_HEX, size=10)

    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # -- data rows (streamed) -------------------------------------------------
    # Amount column indices (1-based): Rate/Night=10, Base=19, Discount=20,
    # Total=21, Refund=22
    AMOUNT_COLS = {10, 19, 20, 21, 22}
    # Date column indices: Booking Date=2, Check-In=11, Check-Out=12,
    # Created=26, Updated=27
    DATE_COLS   = {2, 11, 12, 26, 27}

    row_count = 0
    col_widths: dict[int, int] = {
        i: len(h) for i, h in enumerate(EXPORT_HEADERS, start=1)
    }
    progress_hint = progress_total_hint or 0
    last_reported_pct = -1

    for booking in qs.iterator(chunk_size=1000):
        row_count += 1
        row_data = _booking_to_row(booking)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_count + 1, column=col_idx, value=value)

            if col_idx in AMOUNT_COLS:
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in DATE_COLS:
                cell.alignment = Alignment(horizontal="center")

            # Track max content width for auto-fit
            col_widths[col_idx] = max(col_widths[col_idx], len(str(value)))

        if on_progress and progress_hint > 0 and row_count % 100 == 0:
            pct = min(95, 15 + int(80 * row_count / progress_hint))
            if pct != last_reported_pct:
                last_reported_pct = pct
                on_progress(pct, row_count)

    # -- auto column widths ---------------------------------------------------
    # Cap at 60 chars to avoid excessively wide columns (e.g. long notes).
    for col_idx, width in col_widths.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 60)

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(file_path))

    logger.info(
        "Bookings xlsx written: path=%s rows=%d",
        file_path,
        row_count,
    )
    return row_count


# ---------------------------------------------------------------------------
# Main entry point (called from Celery task)
# ---------------------------------------------------------------------------


def run_booking_export(
    *,
    export_id: str,
    requesting_user_id,
) -> dict[str, Any]:
    """Orchestrate a complete booking export job.

    Called by ``bookings.tasks.generate_booking_export``.

    Steps
    -----
    1. Load ``BookingExport`` record, set status → PROCESSING.
    2. Reload requesting user for branch-scope check.
    3. Rebuild queryset from ``filters_applied`` snapshot.
    4. Write xlsx via ``build_bookings_xlsx``.
    5. Update record: status → READY, file paths, record_count, expires_at.

    Returns a dict with result metadata for the Celery result backend.
    """
    from django.contrib.auth import get_user_model

    from bookings.models import BookingExport

    User = get_user_model()

    export = BookingExport.objects.select_related("requested_by", "branch").get(pk=export_id)
    export.status           = BookingExport.Status.PROCESSING
    export.progress_percent = 10
    export.export_started_at = timezone.now()
    export.save(update_fields=["status", "progress_percent", "export_started_at", "updated_at"])

    try:
        requesting_user = export.requested_by
        filters         = export.filters_applied

        # Build scoped queryset — select_related prevents N+1
        qs = build_booking_export_queryset(filters, requesting_user)
        progress_total = export.estimated_count or qs.count()
        if export.estimated_count is None:
            export.estimated_count = progress_total
            export.save(update_fields=["estimated_count", "updated_at"])

        export.progress_percent = 15
        export.save(update_fields=["progress_percent", "updated_at"])

        # Resolve output path
        export_dir = Path(getattr(settings, "BOOKING_EXPORT_DIR",
                                  Path(settings.MEDIA_ROOT) / "exports" / "bookings"))
        export_dir.mkdir(parents=True, exist_ok=True)

        timestamp  = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename   = f"bookings_export_{export_id[:8]}_{timestamp}.xlsx"
        file_path  = export_dir / filename

        def _report_progress(pct: int, _rows: int) -> None:
            BookingExport.objects.filter(pk=export_id).update(progress_percent=pct)

        record_count = build_bookings_xlsx(
            qs,
            file_path,
            on_progress=_report_progress,
            progress_total_hint=progress_total,
        )

        download_url   = booking_export_download_api_path(str(export_id))
        retention_days = getattr(settings, "BOOKING_EXPORT_RETENTION_DAYS", 7)
        expires_at     = timezone.now() + timedelta(days=retention_days)

        export.status             = BookingExport.Status.READY
        export.file_path          = str(file_path)
        export.download_url       = download_url
        export.record_count       = record_count
        export.progress_percent   = 100
        export.expires_at         = expires_at
        export.export_finished_at = timezone.now()
        export.save(update_fields=[
            "status", "file_path", "download_url", "record_count", "progress_percent",
            "expires_at", "export_finished_at", "updated_at",
        ])

        logger.info(
            "Booking export READY: id=%s rows=%d user=%s role=%s",
            export_id,
            record_count,
            requesting_user.phone,
            requesting_user.role,
        )

        return {
            "export_id":     str(export_id),
            "record_count":  record_count,
            "download_url":  download_url,
            "filename":      filename,
            "generated_at":  datetime.now().isoformat(),
        }

    except Exception as exc:
        export.status        = BookingExport.Status.FAILED
        export.error_message = str(exc)
        export.export_finished_at = timezone.now()
        export.save(update_fields=[
            "status", "error_message", "export_finished_at", "updated_at",
        ])
        logger.exception("Booking export FAILED: id=%s error=%s", export_id, exc)
        raise
